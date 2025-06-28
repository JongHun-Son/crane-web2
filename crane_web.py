import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime
import os

# ✅ 앱 실행 위치 기준 절대 경로 사용
BASE_PATH = os.path.dirname(os.path.abspath(__file__))  # 현재 .py 파일 위치
SAVE_FOLDER = os.path.join(BASE_PATH, "계획서기록")  # ./계획서기록
os.makedirs(SAVE_FOLDER, exist_ok=True)

# ✅ Streamlit 페이지 설정
st.set_page_config(page_title="중량물 작업계획서", page_icon="📝", layout="centered")
st.title("📝 중량물 작업계획서 자동 생성기")
st.markdown("필요한 정보를 입력하고 엑셀 파일을 생성하세요.")

# ✅ 입력 항목
fields = {
    "작업명": "",
    "작업일자": datetime.today().strftime('%Y-%m-%d'),
    "작업 위치": "",
    "중량물 명칭": "",
    "중량": "",
    "크기": "",
    "사용 장비": "",
    "작업 책임자": "",
    "위험요소": "",
    "안전조치": ""
}
data = {}

# ✅ 입력 폼
with st.form("form"):
    for key, default in fields.items():
        data[key] = st.text_input(key, value=default)
    submitted = st.form_submit_button("📁 엑셀 파일 생성")

# ✅ 엑셀 파일 생성 및 저장
if submitted:
    # 필수 필드 검사 (작업명, 작업일자)
    if not data.get("작업명"):
        st.error("❗ '작업명'은 반드시 입력해야 합니다.")
        st.stop()

    wb = Workbook()
    ws = wb.active
    ws.title = "중량물 작업계획서"

    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True)

    # 제목
    ws.merge_cells("A1:B1")
    ws["A1"] = "중량물 작업계획서"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center

    # 내용
    row = 3
    for key, value in data.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].alignment = align_center
        ws[f"B{row}"].alignment = align_wrap
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    # ✅ 파일명 안전하게 생성
    safe_name = data.get('작업명', '작업').replace(" ", "_") or "작업"
    safe_date = data.get('작업일자', datetime.today().strftime('%Y-%m-%d'))
    filename = f"{safe_date}_{safe_name}.xlsx"
    filepath = os.path.join(SAVE_FOLDER, filename)

    # ✅ 엑셀 파일 로컬 저장
    wb.save(filepath)

    # ✅ Streamlit 다운로드용 메모리 버퍼 생성
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ✅ 사용자 알림 및 다운로드 버튼
    st.success(f"✅ 엑셀 파일 생성 완료!\n\n📁 저장 위치: `{filepath}`")
    st.download_button(
        label="📥 엑셀 파일 다운로드",
        data=buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
